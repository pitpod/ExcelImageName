# -*- coding: utf-8 -*-
import os
import sys
import configparser
import math
import errno
import shutil
import pandas as pd
import openpyxl
import subprocess
import glob
from PIL import Image
# from openpyxl.drawing.image import Image
from openpyxl.styles  import Font
from PyQt5 import QtWidgets, QtGui, QtCore
from PyQt5.QtGui import QIcon, QPaintDevice, QPixmap
from PyQt5.QtCore import Qt, QTimer, QAbstractTableModel
from PyQt5.QtWidgets import QMainWindow, QLineEdit, QFileDialog, QMessageBox, QApplication, QGraphicsScene, QSizePolicy, QGraphicsPixmapItem
from configobj import ConfigObj
from Ui_image_excel import Ui_MainWindow

class Application(QMainWindow):
    def __init__(self, parent=None):
        super(Application, self).__init__(parent)
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        self.view = self.ui.graphicsView
        self.scene = QGraphicsScene()
        self.ui.graphicsView.setScene(self.scene)
        ini_cur_path = os.path.dirname(__file__)
        self.config_ini = configparser.ConfigParser()
        self.config_ini_path = f'{ini_cur_path}/config.ini'
        self.name_image_columns = self.config_read()
        self.abc =["A","B","C","D","E","F","G","H","I","J","K","L","M","N","O","P","Q","R","S","T","U","V","W","X","Y","Z"]
        unit = ["%","px"]
        self.ui.comboBox_2.addItems(unit)
        self.ui.lineEdit_5.setText(self.name_image_columns[0])
        self.ui.lineEdit_6.setText(self.name_image_columns[1])
        self.ui.lineEdit_2.setText(self.name_image_columns[2])
        self.ui.lineEdit_3.setText(self.name_image_columns[3])
        self.ui.lineEdit_4.setText(self.name_image_columns[4])
        self.ui.lineEdit_7.setText("3")
        self.ui.lineEdit_9.setText(self.name_image_columns[5])
        self.ui.pushButton.clicked.connect(lambda: self.excel_select())
        self.ui.pushButton_2.clicked.connect(lambda: self.insert_image())
        self.ui.pushButton_3.clicked.connect(lambda: self.openFiles())
        self.ui.pushButton_4.clicked.connect(lambda: self.excel_read())
        self.ui.pushButton_5.clicked.connect(lambda: self.file_rename())
        self.ui.pushButton_6.clicked.connect(lambda: self.excel_open())
        self.ui.tableView_2.clicked.connect(lambda: self.select_file_node(self.ui.tableView_2.currentIndex()))
        self.ui.tableView.setContextMenuPolicy(QtCore.Qt.CustomContextMenu)
        self.ui.tableView.customContextMenuRequested.connect(self.contextMenu)
        self.ui.tableView_2.setContextMenuPolicy(QtCore.Qt.CustomContextMenu)
        self.ui.tableView_2.customContextMenuRequested.connect(self.contextMenu_2)
        self.ui.pushButton_8.clicked.connect(lambda: self.startRowCopy())
        QTimer.singleShot(1, self.imageView)
        headders = ['画像ファイル名','ファイルサイズ']
        clear_list = [["",""]]
        clear_df = pd.DataFrame(clear_list)
        self.clear_model = IM_Model(clear_df, headders) 
        self.ui.tableView_2.setModel(self.clear_model)


    def config_read(self):
        if os.path.exists(self.config_ini_path):
            with open(self.config_ini_path, encoding='utf-8') as fp:
                self.config_ini.read_file(fp)
                conf = self.config_ini['IMAGE']
                type_cell = conf.get('type_column')
                folder_cell = conf.get('folder_column')
                name_cell = conf.get('file_column')
                image_cell = conf.get('image_column')
                image_size = conf.get('image_size')
                book_no = conf.get('book_no')
            return type_cell, folder_cell, name_cell, image_cell, image_size, book_no
        else:
            raise FileNotFoundError(errno.ENOENT, os.strerror(errno.ENOENT), self.config_ini_path)

    def config_path(self):
        if os.path.exists(self.config_ini_path):
            with open(self.config_ini_path, encoding='utf-8') as fp:
                self.config_ini.read_file(fp)
                path = self.config_ini['PATH']
                cur_path = path.get('cur_path')
            return cur_path
        else:
            raise FileNotFoundError(errno.ENOENT, os.strerror(errno.ENOENT), self.config_ini_path)

    def excel_open(self):
        open_path = self.ui.lineEdit.text()
        active_sheet = self.ui.comboBox.currentText()
        """
        workbook = openpyxl.load_workbook(open_path)
        workbook.active = workbook[active_sheet]
        workbook.save(open_path)
        workbook.close()
        """
        subprocess.Popen(['open',"-a", "Microsoft Excel", open_path])

    def excel_select(self):
        fname = QFileDialog.getOpenFileName(self, 'Open file', os.path.expanduser('~') + '/Desktop')
        filepath = fname[0]
        if filepath == "":
            return "break"
        self.ui.listWidget.clear()
        self.ui.comboBox.clear()
        self.ui.lineEdit.setText(filepath)
        self.wb = openpyxl.load_workbook(filepath)
        active_sheet_name = self.wb.active.title
        self.wsn = active_sheet_name
        sheets = self.wb.sheetnames
        self.ui.comboBox.addItems(sheets)
        self.ui.comboBox.setCurrentText(active_sheet_name)
        self.wb.close()


    def startRowCopy(self):

        nextStart = self.ui.lineEdit_8.text()
        self.ui.lineEdit_7.setText(nextStart)
        self.ui.lineEdit_8.clear()

    def excel_read(self):
        # im_model = IM_Model() 
        self.ui.tableView_2.setModel(self.clear_model)
        self.ui.listWidget.clear()
        self.scene.clear()
        """
        sk_list = []
        for i in range(3, sk):
            sk_list.append(i)
        """
        name = self.ui.lineEdit.text()
        sh = self.ui.comboBox.currentText()
        type_col = self.abc.index(self.name_image_columns[0])
        folder_col = self.abc.index(self.name_image_columns[1])
        file_col = self.abc.index(self.name_image_columns[2])
        img_col = self.abc.index(self.name_image_columns[3])
        bookno_col = self.abc.index(self.name_image_columns[5])
        use_cols = [type_col, bookno_col, folder_col, file_col, img_col]

        ws = self.wb[sh]
        maxRow = ws.max_row + 1
        for i in reversed(range(1, maxRow)):
            if ws.cell(row=i, column=img_col + 1).value != None:
                last = i
                break
        bkn = self.ui.lineEdit_10.text()
        if bkn == "":
            pass
        else:
            bkncell = self.ui.lineEdit_9.text()
            # df = df[df['ブックNo'] == bkn]
            # ws = self.wb[self.wsn]
            ret = self.search_column(ws[bkncell], bkn)
            srow = ret[0]
            srow = srow.replace(bkncell,"")
            srow = int(srow) - 1
            erow = ret[-1]
            erow = erow.replace(bkncell,"")
            erow = int(erow)
            self.ui.lineEdit_7.setText(str(srow))
            self.ui.lineEdit_8.setText(str(erow))

        sk = int(self.ui.lineEdit_7.text())
        ft = self.ui.lineEdit_8.text()
        if ft == "":
            foot = 0
        else:
            foot = int(maxRow) - int(ft) - 1
        # headers = ['種別','フォルダ名','ファイル名','画像ファイル名']
        headers = ['種別', 'ブックNo','フォルダ名','ファイル名','画像ファイル名']
        df = pd.read_excel(name, sheet_name=sh, dtype=str, header=None, names=headers, usecols=use_cols, skiprows=sk, skipfooter=foot)
        serila_list = []
        if self.ui.checkBox_2.isChecked():
            for num in range(len(ret)):
                string = str(num + 1).zfill(3)
                serila_list.append(string)
            # serial_num = pd.RangeIndex(start=1, stop=len(ret) + 1, step=1).astype(str)
            # serial_num = serial_num.to_list()
            # serial_num = serial_num.astype(str)
            # 各セルに対して関数を適用
            width = 4  # 揃えたい桁数
            # formatted_df = serial_num.applymap(lambda x: self.format_cell(x, width))
            df['画像ファイル名'] = serila_list
        self.df = df.dropna(subset=['画像ファイル名'])
        # 重複チェック
        dup = self.df[self.df.duplicated(subset='画像ファイル名', keep='first')]
        if dup.empty == False:
            status_text = ""
            for row in dup.itertuples():
                status_text = f'{row[3]}---{row[4]}が重複しています'
                self.ui.listWidget.addItem(status_text)
            self.ui.listWidget.addItem("-----------------------------------")
            # self.re_model.sort('画像名', True)
            return "break"
        self.np_model = IE_Model(self.df, headers) 
        self.ui.tableView.setModel(self.np_model)
        self.ui.tableView.setColumnWidth(0, 40)
        self.ui.tableView.setColumnWidth(1, 60)
        self.ui.tableView.setColumnWidth(2, 300)
        self.ui.tableView.setColumnWidth(3, 360)
        self.ui.tableView.setColumnWidth(4, 80)
    # 各セルの値を指定の桁数に揃える関数を定義
    def format_cell(self, value, width):
        return f"{value:0{width}}"
    # 特定の列を検索
    def search_column(self, column, keyword):
        result = []
        for cell in column:
            # セルのデータを文字列に変換
            try:
                value = str(cell.value)
            # 文字列に変換できないデータはスキップ
            except:
                continue
            # キーワードに一致するセルの番地を取得
            if value == keyword:
                cell_address = openpyxl.utils.get_column_letter(cell.column) +  str(cell.row)
                result.append(cell_address)
    
        return result
    def openFiles(self, select_type = 0):
        self.scene.clear()
        cur_path = self.config_path()
        if cur_path == "null":
            cur_path = os.path.expanduser('~') + '/Desktop'

        dir_path = QFileDialog.getExistingDirectory(self, 'Open Directory', cur_path)
        if dir_path == "":
            return "break"
        else:
            new_path = os.path.dirname(dir_path)
            config = ConfigObj(self.config_ini_path, encoding='utf-8')
            config['PATH']['cur_path'] = new_path
            config.write()
        self.ui.label_7.setText(dir_path)
        fileNames = glob.glob(f'{dir_path}/*')
        self.image_list = fileNames
        flList = []
        self.fnames = []
        if 0 < len(fileNames):
            for name in fileNames:
                ext = os.path.splitext(os.path.basename(name))[1]
                if ext != ".xlsx":
                    fSize = self.convert_size(os.path.getsize(name), 'MB') 
                    fname = os.path.splitext(os.path.basename(name))[0]
                    flList.append([name, fname, fSize])
                    self.fnames.append(os.path.basename(name))
            column_list = ['ファイルパス', '画像ファイル名','ファイルサイズ']
            """
            flList_df = pd.DataFrame(flList, columns=column_list)
            flList_df['画像ファイル名'] = flList_df['画像ファイル名'].astype(str).str.zfill(1)
            self.flList_df = flList_df.sort_values('画像ファイル名')
            """

            self.flList_df = pd.DataFrame(flList, columns=column_list).sort_values('画像ファイル名')
            fl_df = self.flList_df.iloc[:,1:]
            headders = ['画像ファイル名','ファイルサイズ']
            self.im_model = IM_Model(fl_df, headders) 
            self.ui.tableView_2.setModel(self.im_model)
            self.ui.tableView_2.setColumnWidth(0, 100)
        else:
            pass
        size_text = ""
        for index, row in self.flList_df.iterrows():
            mb = float(row[2].replace(" MB",""))
            if mb < float(3.5):
                size_text = f'{row[1]}が3.5メガ未満です。'
                self.ui.listWidget.addItem(size_text)
        if size_text != "":
            return "break"
        # df_list = pd.merge(self.df, self.flList_df.drop('ファイルサイズ', axis=1), on='画像ファイル名', how='left')
        df_list = pd.merge(self.df, self.flList_df.drop('ファイルサイズ', axis=1), on='画像ファイル名', how='outer')
        los_text = ""
        self.ui.listWidget.clear()
        for index, row in df_list.iterrows():
            if str(row[4]) == 'nan':
                los_text = f'{row[4]}が画像フォルダにありません。'
                self.ui.listWidget.addItem(los_text)
            if str(row[2]) == 'nan':
                los_text = f'{row[4]}がリストにありません。'
                self.ui.listWidget.addItem(los_text)
        if los_text != "":
            return "break"

    def select_file_node(self, index):
        index_col = index.column()
        index_row = index.row()

        path_text = self.flList_df.iloc[index_row,0] 
        self.imageView(path_text)

    def imageView(self, f_path = ""):
        f = f_path
        self.scene.clear()
        ixmap = QPixmap(f)
        self.ixmap_height = ixmap.height()
        self.ixmap_width = ixmap.width()
        self.pic_Item = QGraphicsPixmapItem(ixmap)
        self.scene.addItem(self.pic_Item)
        self.scene.setSceneRect(0.0,0.0,self.ixmap_width,self.ixmap_height)
        self.ui.graphicsView.fitInView(self.scene.sceneRect(), Qt.KeepAspectRatio)

    def convert_size(self, size, unit="B"):
        units = ("B", "KB", "MB", "GB", "TB", "PB", "EB", "ZB")
        i = units.index(unit.upper())
        size = round(size / 1024 ** i, 2)

        return f"{size} {units[i]}"

    def file_rename(self):
        # mogrify -path ../rev -resize 50% -quality 100 ./
        resize = int(self.ui.lineEdit_4.text())
        df_rename = pd.merge(self.df, self.flList_df.drop('ファイルサイズ', axis=1), on='画像ファイル名', how='right')
        los_text = ""
        self.ui.listWidget.clear()
        for index, row in df_rename.iterrows():
            if str(row[3]) == 'nan':
                los_text = f'{row[4]}がリストにありません。'
                self.ui.listWidget.addItem(los_text)
        if los_text != "":
            return "break"
        flList_cunt = len(self.flList_df)
        df_count = len(df_rename)
        dir_path = QFileDialog.getExistingDirectory(self, 'Select Folder', os.path.expanduser('~') + '/Desktop')
        type_text = self.ui.lineEdit_11.text()
        book_no = self.ui.lineEdit_10.text()

        if self.ui.checkBox.isChecked():
            dir_path = f'{dir_path}/{type_text}_{book_no}'
        if dir_path == "":
            return "break"
        dup_no = 0
        for column_name, item in df_rename.iterrows():
            origin_path = f'{item[5]}'
            item_1 = item[2].replace('/', '／')
            if type_text != "":
                item_0 = type_text
            else:
                item_0 = item[0]
            folder_name = f'{dir_path}/{item_0}/{item_1}/'
            folder_name = folder_name.replace(':','：')
            if os.path.exists(folder_name) == False:
                os.makedirs(folder_name)
            item_2 = item[3].replace('/', '／')
            rename_path = f'{dir_path}/{item_0}/{item_1}/{item_2}'
            rename_path = rename_path.replace(':','：')

            if os.path.isfile(rename_path):
                los_text = f'{os.path.basename(rename_path)}はすでにあります。'
                self.ui.listWidget.addItem(los_text)
                """
                dup_no += 1
                root, ext = os.path.splitext(rename_path)
                rename_path = f'{root}_重複{dup_no}{ext}'
                """
            else:
                if resize != 100:
                    if self.ui.comboBox_2.currentText() == "%":
                        f_resize = f'{resize}%'
                    else:
                        f_resize = f'{resize}x{resize}'
                    r = subprocess.run(['convert', f'{origin_path}', '-resize', f_resize, f'{rename_path}'], stdout=subprocess.PIPE)
                else:
                    shutil.copy2(origin_path, rename_path)
        ms_text = "終了しました"
        msgBox = QMessageBox()
        msgBox.setText(ms_text)
        msgBox.setIcon(QMessageBox.Icon.Information)
        msgBox.setStandardButtons(QMessageBox.Ok)
        msgBox.exec_()

    def insert_image(self):
        name = self.ui.lineEdit.text()
        ExcelName = self.ui.lineEdit.text()
        wb = openpyxl.load_workbook(name)
        ws = wb.active
        img_position = self.ui.lineEdit_3.text()
        #最大行
        maxRow =  ws.max_row + 1
        #画像を選択 & 挿入
        for f in range(1, maxRow):
            fname = ws.cell(row=f, column=3).value
            if fname != None:
                for fpath in self.image_list:
                    if fname in os.path.basename(fpath):
                        img = Image(fpath)
                        resize = int(self.ui.lineEdit_4.text())
                        col = self.abc.index(img_position) + 1
                        font = ws.cell(row = f, column = col).font
                        # font_size = font.size #フォントサイズ　ピクセル
                        # dpi = QPaintDevice.physicalDpiY(self)
                        # fs = int(font_size * dpi / 72) #ピクセル
                        # fs = int(font_size * 72 / 72) #ピクセル
                        point_size = (resize * 72) / 72 
                        ws.row_dimensions[f].height = str(math.ceil(resize * 0.78))
                        ws.column_dimensions[img_position].width = str(math.ceil(resize * 0.151))
                        if img.width > img.height:
                            img = self.scale_to_width(img, resize)
                        elif img.width < img.height:
                            img = self.scale_to_width(img, resize, 1)
                        elif img.width == img.height:
                            img.width = resize
                            img.height = resize
                        ws.add_image(img, f'{img_position}{f}')

        #保存
        spath = os.path.dirname(ExcelName)
        sname = os.path.basename(ExcelName)
        save_path = f'{spath}/image_{sname}'
        wb.save(save_path)
        QMessageBox.information(None, "通知", "ファイルを書き出しました。", QMessageBox.Yes)

    def scale_to_width(self, img, resize, wh = 0):  # アスペクト比を固定して、幅が指定した値になるようリサイズする。
        if wh == 0:
            width = resize
            height = round(img.height * width / img.width)
            img.width = width
            img.height = height
            return img
        elif wh == 1:
            height = resize
            width = round(img.width * height / img.height)
            img.width = width
            img.height = height
            return img
    
    def delItem(self, view, model, list_type = 0):
        indexes = view.selectedIndexes()
        
        if model.rowCount() == 0:
            return
    
        if len(indexes) == 0:
            model.removeItem(model.rowCount()-1)
            return
        
        rows = set([index.row() for index in indexes])
        if list_type == 0:
            self.df = model.removeItems(rows)
        else:
            self.fl_df = model.removeItems(rows)
            for row in list(rows)[::-1]:
                self.flList_df = self.flList_df.drop([row])
            self.flList_df = self.flList_df.reset_index(drop=True)
            # self.flList_df = self.flList_df.iloc[]

    def contextMenu(self, point):
        self.menu = QtWidgets.QMenu(self)
        # self.menu.addAction('Insert', self.insertRow)
        self.menu.addAction('', self.delItem)
        self.menu.addAction('Delete', lambda:self.delItem(self.ui.tableView, self.np_model))
        self.menu.exec_(self.focusWidget().mapToGlobal(point))

    def contextMenu_2(self, point):
        self.menu = QtWidgets.QMenu(self)
        # self.menu.addAction('Insert', self.insertRow)
        self.menu.addAction('', self.delItem)
        self.menu.addAction('Delete', lambda:self.delItem(self.ui.tableView_2, self.im_model, 1))
        self.menu.exec_(self.focusWidget().mapToGlobal(point))

class MyLineEdit(QLineEdit):
    def mouseDoubleClickEvent(self, e):
        super().mouseDoubleClickEvent(e)
        point_x = e.x()
        point_y = e.y()

        idx = self.cursorPositionAt(e.pos())
        word = self.text()
        """
        start = 0
        end = len(word)

        for n in [i for i,c in enumerate(word) if c in "_ "]:
            if n >= idx:
                end = n
                break
            if n < idx:
                start = n+1

        self.setSelection(start, end-start)
        """

class IM_Model(QAbstractTableModel):
    def __init__(self, list, headers = [], rows = [], parent = None):
        QAbstractTableModel.__init__(self, parent)
        self.list = list
        self.headers = headers
        # self.rows = rows
        self.db_list = []
        self.items = []

    def rowCount(self, parent = None):
        return len(self.list)

    def columnCount(self, parent = None):
        return len(self.list.columns)

    def flags(self, index):
        return QtCore.Qt.ItemIsEditable | QtCore.Qt.ItemIsEnabled | QtCore.Qt.ItemIsSelectable

    def data(self, index, role):
        row = index.row()
        column = index.column()
        value = self.list.iat[row, column]

        if role == Qt.EditRole:
            value = self.list.iat[row, column]
            return value

        if role == Qt.DisplayRole:
            row = index.row()
            column = index.column()
            value = self.list.iat[row, column]
            return value

    def headerData(self, section, orientation, role):
        if role == Qt.DisplayRole:
            if orientation == Qt.Horizontal:
                if section < len(self.headers):
                    return self.headers[section]
                else:
                    return "not implemented"
            else:
                return f'{section + 1}'

    def removeItem(self, row, parent=QtCore.QModelIndex()):
        self.beginRemoveRows(parent, row, row)
        # del self.list[row]
        self.endRemoveRows()

    def removeItems(self, rows):
        self.list = self.list.reset_index(drop=True)
        for row in list(rows)[::-1]:
            # row = row + 1
            self.beginRemoveRows(QtCore.QModelIndex(), row, row)
            # del self.list[row]
            self.list = self.list.drop([row])
            self.endRemoveRows() 
        self.list = self.list.reset_index(drop=True)
        return self.list
    
    def addItem(self, row, item, parent=QtCore.QModelIndex()):
        self.beginInsertRows(parent, row, row)
        self.list.insert(row, item)
        self.endInsertRows()

class IE_Model(QAbstractTableModel):
    def __init__(self, list, headers = [], rows = [], parent = None):
        QAbstractTableModel.__init__(self, parent)
        self.list = list
        self.headers = headers
        # self.rows = rows
        self.db_list = []
        self.items = []

    def rowCount(self, parent = None):
        return len(self.list)

    def columnCount(self, parent = None):
        return len(self.list.columns)

    def flags(self, index):
        return QtCore.Qt.ItemIsEditable | QtCore.Qt.ItemIsEnabled | QtCore.Qt.ItemIsSelectable

    def data(self, index, role):
        row = index.row()
        column = index.column()
        value = self.list.iat[row, column]

        if role == Qt.EditRole:
            value = self.list.iat[row, column]
            return value

        if role == Qt.DisplayRole:
            row = index.row()
            column = index.column()
            value = self.list.iat[row, column]
            return value

    def headerData(self, section, orientation, role):
        if role == Qt.DisplayRole:
            if orientation == Qt.Horizontal:
                if section < len(self.headers):
                    return self.headers[section]
                else:
                    return "not implemented"
            else:
                return f'{section + 1}'

    def removeItem(self, row, parent=QtCore.QModelIndex()):
        self.beginRemoveRows(parent, row, row)
        # del self.list[row]
        self.endRemoveRows()

    def removeItems(self, rows):
        self.list = self.list.reset_index(drop=True)
        for row in list(rows)[::-1]:
            # row = row + 1
            self.beginRemoveRows(QtCore.QModelIndex(), row, row)
            # del self.list[row]
            self.list = self.list.drop([row])
            self.endRemoveRows() 
        self.list = self.list.reset_index(drop=True)
        return self.list
    
    def addItem(self, row, item, parent=QtCore.QModelIndex()):
        self.beginInsertRows(parent, row, row)
        self.list.insert(row, item)
        self.endInsertRows()

class Delegate(QtWidgets.QStyledItemDelegate):
    def __init__(self, parent=None, setModelDataEvent=None):
        super(Delegate, self).__init__(parent)
        self.setModelDataEvent = setModelDataEvent

    def createEditor(self, parent, option, index):
        return QtWidgets.QLineEdit(parent)

    def setEditorData(self, editor, index):
        value = index.model().data(index, QtCore.Qt.DisplayRole)
        editor.setText(str(value))

    def setModelData(self, editor, model, index):
        model.setData(index, editor.text())
        if not self.setModelDataEvent is None:
            self.setModelDataEvent()

def resource_path(relative):
    if hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, relative)
    return os.path.join(os.path.abspath('.'), relative)

def main():
    app = QApplication(sys.argv)
    app.setWindowIcon(QIcon(resource_path('image/re.png')))
    app.setStyle(QtWidgets.QStyleFactory.create('Fusion')) # won't work on windows style.
    main_app = Application(None)
    main_app.show()
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()